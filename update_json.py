import os
import json
import re
import requests
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from pathlib import Path

# ===============================
# CONFIG
# ===============================
EXTRACT_ROOT = "extracted"
OUTPUT_EXCEL = "recipes_report.xlsx"

SMARTSHEET_TOKEN = "7xcmOm3neR6SXBXda7fY9qis3Bg9z9VsBZ6T6"
SMARTSHEET_SHEET_ID = "7220178429366148"

SMARTSHEET_HEADERS = {
    "Authorization": f"Bearer {SMARTSHEET_TOKEN}",
    "Content-Type": "application/json"
}

# ===============================
# HELPERS
# ===============================

def safe_int(val, default=0):
    try:
        return int(val)
    except:
        return default


def clean_line(text):
    return text.replace(":", "").replace("-", "").strip()


def normalize_minutes(text):
    text = text.upper().strip()
    match = re.search(r"(\d+)", text)
    if not match:
        return ""
    return f"{match.group(1)} mins"


# ===============================
# SMARTSHEET LOGIC
# ===============================

def load_smartsheet_data():
    print("📡 Fetching Smartsheet data...")

    url = f"https://api.smartsheet.com/2.0/sheets/{SMARTSHEET_SHEET_ID}"
    response = requests.get(url, headers=SMARTSHEET_HEADERS)

    if response.status_code != 200:
        raise Exception(f"Smartsheet API error: {response.status_code} - {response.text}")

    sheet = response.json()

    column_map = {}
    for col in sheet["columns"]:
        column_map[col["id"]] = col["title"]

    lookup = {}

    for row in sheet["rows"]:
        row_data = {}
        recipe_name = None

        for cell in row["cells"]:
            col_name = column_map.get(cell["columnId"])
            value = cell.get("value", "")

            if col_name == "Recipe Name":
                recipe_name = str(value).strip().upper()
            elif col_name in ["Veg/Non Veg", "Cooking Mode", "Cuisine", "Category"]:
                row_data[col_name] = str(value).strip().upper()

        if recipe_name:
            lookup[recipe_name] = row_data

    print(f"✅ Loaded {len(lookup)} recipes from Smartsheet")
    return lookup


# ===============================
# TXT PARSER
# ===============================

def parse_recipe_txt(txt_path):
    with open(txt_path, "r", encoding="utf-8") as f:
        data = json.load(f)

    recipe_name = data.get("name", [""])[0].strip().upper()

    total_on2cook_time_sec = sum(safe_int(step.get("durationInSec", 0)) for step in data.get("Instruction", []))
    total_on2cook_time_min = max(1, round(total_on2cook_time_sec / 60))

    description = data.get("description", "")
    desc_upper = description.upper()

    # Final Output
    total_output = ""
    if "FINAL OUTPUT" in desc_upper:
        try:
            total_output = clean_line(desc_upper.split("FINAL OUTPUT")[1].split("\n")[0])
        except:
            pass
    elif "OUTPUT" in desc_upper:
        try:
            total_output = clean_line(desc_upper.split("OUTPUT")[1].split("\n")[0])
        except:
            pass

    # Accessories
    accessories = ""
    if "ACCESSORIES" in description.upper():
        try:
            part = description.split("ACCESSORIES", 1)[1]
            acc_list = []

            for line in part.splitlines():
                line = line.strip()
                if not line:
                    continue

                upper = line.upper()

                stop_keywords = [
                    "FINAL OUTPUT", "OUTPUT", "NORMAL COOKING TIME", "NORMAL TIME",
                    "OTHER ESSENTIALS", "INGREDIENTS", "NOTE", "SPECIAL INSTRUCTION"
                ]
                
                if any(keyword in upper for keyword in stop_keywords):
                    break

                acc_list.append(line.title())

            accessories = ", ".join(acc_list)

        except Exception as e:
            print(f"❌ Accessories parse error: {e}")

    # Normal Cooking Time
    normal_cooking_time = ""
    if "NORMAL COOKING TIME" in desc_upper:
        try:
            raw = clean_line(desc_upper.split("NORMAL COOKING TIME")[1].split("\n")[0])
            normal_cooking_time = normalize_minutes(raw)
        except:
            pass
    elif "NORMAL TIME" in desc_upper:
        try:
            raw = clean_line(desc_upper.split("NORMAL TIME")[1].split("\n")[0])
            normal_cooking_time = normalize_minutes(raw)
        except:
            pass

    recipe = {
        "Recipe Name": recipe_name,
        "Veg/Non Veg": "VEG",
        "Cooking Mode": "AUTO",
        "Cuisine": "GLOBAL CUISINE",
        "Category": "MAIN COURSE",
        "Accessories": accessories,
        "Total Output": total_output,
        "On2Cook Cooking Time": f"{total_on2cook_time_min} mins",
        "Normal Cooking Time": normal_cooking_time
    }

    return recipe


# ===============================
# EXCEL GENERATOR
# ===============================

def generate_excel():
    recipes_data = []
    
    smartsheet_map = load_smartsheet_data()

    for recipe_key in os.listdir(EXTRACT_ROOT):
        recipe_dir = os.path.join(EXTRACT_ROOT, recipe_key)
        if not os.path.isdir(recipe_dir):
            continue

        txt_file = next((f for f in os.listdir(recipe_dir) if f.lower().endswith(".txt")), None)
        if not txt_file:
            print(f"⚠ No TXT file in {recipe_key}")
            continue

        recipe = parse_recipe_txt(os.path.join(recipe_dir, txt_file))
        name = recipe["Recipe Name"]

        # Smartsheet override
        if name in smartsheet_map:
            ss = smartsheet_map[name]
            for field in ["Veg/Non Veg", "Cooking Mode", "Cuisine", "Category"]:
                old = recipe[field]
                new = ss.get(field, old)
                if new and new != old:
                    recipe[field] = new
                    print(f"🔄 {name} → {field}: {old} → {new}")
        else:
            print(f"⚠ {name} not found in Smartsheet")

        recipes_data.append({
            "Zip File Name": recipe_key,
            "Recipe Name": name,
            "Accessories": recipe["Accessories"],
            "Output": recipe["Total Output"],
            "On2Cook Cooking Time": recipe["On2Cook Cooking Time"],
            "Normal Cooking Time": recipe["Normal Cooking Time"],
            "Veg/Non Veg": recipe["Veg/Non Veg"],
            "Cooking Mode": recipe["Cooking Mode"],
            "Cuisine": recipe["Cuisine"],
            "Category": recipe["Category"]
        })
        
        print(f"🍽 Parsed: {name}")

    # Create Excel file
    wb = Workbook()
    sheet = wb.active
    sheet.title = "Recipes"

    # Headers
    headers = [
        "Zip File Name",
        "Recipe Name",
        "Accessories",
        "Output",
        "On2Cook Cooking Time",
        "Normal Cooking Time",
        "Veg/Non Veg",
        "Cooking Mode",
        "Cuisine",
        "Category"
    ]

    # Style headers
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")

    for col_num, header in enumerate(headers, 1):
        cell = sheet.cell(row=1, column=col_num)
        cell.value = header
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment

    # Add data
    for row_num, recipe in enumerate(recipes_data, 2):
        sheet.cell(row=row_num, column=1, value=recipe["Zip File Name"])
        sheet.cell(row=row_num, column=2, value=recipe["Recipe Name"])
        sheet.cell(row=row_num, column=3, value=recipe["Accessories"])
        sheet.cell(row=row_num, column=4, value=recipe["Output"])
        sheet.cell(row=row_num, column=5, value=recipe["On2Cook Cooking Time"])
        sheet.cell(row=row_num, column=6, value=recipe["Normal Cooking Time"])
        sheet.cell(row=row_num, column=7, value=recipe["Veg/Non Veg"])
        sheet.cell(row=row_num, column=8, value=recipe["Cooking Mode"])
        sheet.cell(row=row_num, column=9, value=recipe["Cuisine"])
        sheet.cell(row=row_num, column=10, value=recipe["Category"])

    # Auto-adjust column widths
    for column in sheet.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        sheet.column_dimensions[column_letter].width = adjusted_width

    wb.save(OUTPUT_EXCEL)
    
    print(f"\n✅ {len(recipes_data)} recipes written to {OUTPUT_EXCEL}")


if __name__ == "__main__":
    generate_excel()